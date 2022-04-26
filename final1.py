from tkinter import *

from tkinter import messagebox

from easygui import fileopenbox, filesavebox 

from openpyxl import open

from openpyxl.styles import Font, Alignment, Border, Side

from math import ceil

main_window = Tk()

def locations(operation):
	if operation == 'open':
		global olabel
		global input_location
		if olabel:
			oempty = Label(main_window,text=' '*1000).place(x = 123, y = 13)	#Clear the label if it's second press
		input_location = fileopenbox()
		input_location_label =  Label(main_window, text=input_location).place(x = 123, y = 13)
		olabel = True
	if operation == 'save':
		global ilabel
		global output_location
		if ilabel:
			oempty = Label(main_window,text=' '*1000).place(x = 110, y = 63)	#Clear the label if it's second press
		output_location = filesavebox()
		if '.' not in output_location:
			output_location += '.xlsx'
		output_location_label =  Label(main_window, text=output_location).place(x = 110, y = 63)
		ilabel = True


def main():

	output_example = "example.xlsx"

	openfile_sheet_name = "Факел"

	


	#Обозначение переменных
	global works
	global materials
	global start_input_cell
	global input_location
	global output_location

	openfile_book = open(input_location, read_only=True)

	output_book = open(output_example, read_only=False)

	for i in output_book.sheetnames:
		if "акт" in i:
			output_sheet_name = i
		if "счет" in i:
			output_sheet_name2 = i	

	book = output_book[output_sheet_name]
	book1 = output_book[output_sheet_name2]
	input_book = openfile_book[openfile_sheet_name]

	if isinstance(works,int) == False:
		works = int(works.get())

		materials = int(materials.get())

		start_output_cell = 26

		start_input_cell = int(start_input_cell.get())

	difference = start_output_cell - start_input_cell

	works_sum = 0

	materials_sum = 0

	current_sum = 0



	letters = [["A","A"],["O","B"],["BJ","C"],["BR","D"],["CB","E"],["CL","F"]]

	for row in range(start_output_cell, start_output_cell + works + 1 + 1 + materials + 3):
		if (row < start_output_cell + works) or ((row > start_output_cell + works + 1) and (row < start_output_cell + works + 1 + 1 + materials )) :
			#Объединение клеток
			book.merge_cells(f"A{row}:N{row}")
			book.merge_cells(f"O{row}:AY{row}")
			book.merge_cells(f"AZ{row}:BI{row}")
			book.merge_cells(f"BJ{row}:BQ{row}")
			book.merge_cells(f"BR{row}:BZ{row}")
			book.merge_cells(f"CB{row}:CK{row}")
			book.merge_cells(f"CL{row}:CU{row}")
			#Запись в клетку + параметры клеток

			current_sum = int(input_book[f"E{str(row - difference)}"].value) * int(input_book[f"D{str(row - difference)}"].value)
			if (row < start_output_cell + works):
				works_sum += current_sum
			else:
				materials_sum += current_sum


			for i in range(len(letters)):
				if letters[i][1] == "F":
					book[f"{letters[i][0]}{str(row)}"].value = current_sum
				else:
					book[f"{letters[i][0]}{str(row)}"].value = str(input_book[f"{letters[i][1]}{str(row - difference)}"].value).lstrip()

				book[f"{letters[i][0]}{str(row)}"].font = Font(name='Times New Roman',size=12, bold=False)
				book[f"{letters[i][0]}{str(row)}"].alignment = Alignment(wrap_text=True)
				if letters[i][1] != "B":
					book[f"{letters[i][0]}{str(row)}"].alignment = Alignment(horizontal='center', vertical='center')

			#Параметры строки 
			book.row_dimensions[row].height = 15 * ((len(str(book[f"O{row}"].value)) // 80 ) + 1)
			line = 0

			while True:	
				try:
					book[row][line].border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
				except:
					break
				line += 1
			#print("ну я сделал строку", row)
		if row == start_output_cell + works or row == start_output_cell + works + 1 + 1 + materials:
			book.merge_cells(f"A{row}:N{row}")
			book.merge_cells(f"O{row}:AY{row}")
			book.merge_cells(f"AZ{row}:BQ{row}")
			book.merge_cells(f"BR{row}:CK{row}")
			book.merge_cells(f"CL{row}:CU{row}")
			if row == start_output_cell + works:
				book[f"BR{str(row)}"].value = "Итого работа:"
				book[f"CL{str(row)}"].value = works_sum
			else:
				book[f"BR{str(row)}"].value = "Итого материалы:"
				book[f"CL{str(row)}"].value = materials_sum

			book[f"BR{str(row)}"].font = Font(name='Times New Roman',size=14, bold=True)
			book[f"BR{str(row)}"].alignment = Alignment(horizontal='center', vertical='center')
			book[f"CL{str(row)}"].font = Font(name='Times New Roman',size=14, bold=True)
			book[f"CL{str(row)}"].alignment = Alignment(horizontal='center', vertical='center')
			book.row_dimensions[row].height = 18
			line = 0
			while True:	
				try:
					book[row][line].border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
				except:
					break
				line += 1

		if row == start_output_cell + works + 1:
			book.merge_cells(f"A{row}:N{row}")
			book.merge_cells(f"O{row}:AY{row}")
			book.merge_cells(f"AZ{row}:BQ{row}")
			book.merge_cells(f"BR{row}:BZ{row}")
			book.merge_cells(f"CB{row}:CK{row}")
			book.merge_cells(f"CL{row}:CU{row}")
			book[f"O{str(row)}"].value = "Материалы:"
			book[f"O{str(row)}"].font = Font(name='Times New Roman',size=14, bold=False)
			book[f"O{str(row)}"].alignment = Alignment(horizontal='center', vertical='center')
			book.row_dimensions[row].height = 18
			line = 0
			while True:	
				try:
					book[row][line].border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
				except:
					break
				line += 1

		if row == start_output_cell + works + 1 + 1 + materials + 1:
			book.merge_cells(f"A{row}:N{row}")
			book.merge_cells(f"O{row}:AY{row}")
			book.merge_cells(f"AZ{row}:BQ{row}")
			book.merge_cells(f"BR{row}:BZ{row}")
			book.merge_cells(f"CB{row}:CK{row}")
			book.merge_cells(f"CL{row}:CU{row}")
			book[f"CB{str(row)}"].value = "Сумма НДС"
			book[f"CB{str(row)}"].font = Font(name='Times New Roman',size=14, bold=True)
			book[f"CB{str(row)}"].alignment = Alignment(horizontal='center', vertical='center')
			book.row_dimensions[row].height = 18
			line = 0
			while True:	
				try:
					book[row][line].border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
				except:
					break
				line += 1
		if row == start_output_cell + works + 1 + 1 + materials + 1 + 1:
			book.merge_cells(f"A{row}:N{row}")
			book.merge_cells(f"O{row}:AY{row}")
			book.merge_cells(f"AZ{row}:BQ{row}")
			book.merge_cells(f"BR{row}:BZ{row}")
			book.merge_cells(f"CB{row}:CK{row}")
			book.merge_cells(f"CL{row}:CU{row}")
			book[f"CB{str(row)}"].value = "Всего по акту"
			book[f"CB{str(row)}"].font = Font(name='Times New Roman',size=14, bold=True)
			book[f"CB{str(row)}"].alignment = Alignment(horizontal='center', vertical='center')
			book[f"CL{str(row)}"].value = works_sum + materials_sum
			book[f"CL{str(row)}"].font = Font(name='Times New Roman',size=14, bold=True)
			book[f"CL{str(row)}"].alignment = Alignment(horizontal='center', vertical='center')
			book.row_dimensions[row].height = 18
			line = 0
			while True:	
				try:
					book[row][line].border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
				except:
					break
				line += 1

	#Число в текст

	numbers = {1:{"м":"один", "ж":"одна"}, 2:{"м":"два", "ж":"две"},3:"три", 4:"четыре", 5:"пять", 6:"шесть", 7:"семь", 8:"восемь", 9:"девять", 
	        10:"десять", 11:"одиннадцать", 12:"двенадцать", 13:"тринадцать", 14:"четырнадцать", 15:"пятнадцать", 16:"шестнадцать", 17:"семнадцать", 18:"восемнадцать", 19:"девятнадцать"}

	hundreds = {100:"сто", 200:"двести", 300:"триста", 400:"четыреста", 500:"пятьсот", 600:"шестьсот", 700:"семьсот", 800:"восемьсот", 900:"девятьсот"}

	decades = {20:"двадцать", 30:"тридцать", 40:"сорок", 50:"пятьдесят", 60:"шестьдесят", 70:"семьдесят", 80:"восемьдесят", 90:"девяносто"}

	orders = {1000:{"1":"тысяча", "2-4":"тысячи", "ост":"тысяч"}, 1000000:{"1":"миллион", "2-4":"миллиона", "ост":"миллионов"}}

	rubles = {"1":"рубль", "2-4":"рубля", "ост":"рублей"}

	s = str(works_sum + materials_sum)

	if s[-2] != "1" and s[-1] == "1": ruble_ending = "1"
	elif s[-2] != "1" and s[-1] in "234": ruble_ending = "2-4"
	else: ruble_ending = "ост"

	order = 0
	wording = ""

	s = s.zfill(ceil(len(s) / 3) * 3)

	while s != '':
	    sub = s[-3:]

	    if sub[0] != "0": sub_wording = hundreds[int(sub[0])*100]
	    else: sub_wording = ""

	    sub = sub[1:]

	    ending = "ост"

	    if sub[0] != "1":
	        if sub[0] != "0": sub_wording += " " + decades[int(sub[0])*10]

	        if sub[1] == "1": ending = "1"
	        elif sub[1] in "234": ending = "2-4"

	        if sub[1] in ["1", "2"]:
	            if order == 1: sub_wording += " " + numbers[int(sub[1])]['ж']
	            else: sub_wording += " " + numbers[int(sub[1])]['м']

	        elif sub[1] != "0":
	            sub_wording += " " + numbers[int(sub[1])]

	    else:
	        sub_wording += " " + numbers[int(sub)]

	    if order == 0: wording = sub_wording.lstrip() + " " + wording
	    else: wording = sub_wording.lstrip() + " " + orders[1000**order][ending] + " " + wording

	    order += 1
	    s = s[:-3]

	string = wording.strip(" ") + " " + rubles[ruble_ending]	





	#Перенос в "счет"
	book1["D22"].value = input_book["A10"].value
	book1["AB22"].value = works_sum + materials_sum
	book1["AG22"].value = works_sum + materials_sum
	book1["AG24"].value = works_sum + materials_sum
	book1["AG26"].value = works_sum + materials_sum
	book1["B27"].value = "Всего наименований 1, на сумму " + str(works_sum + materials_sum) + " " + " ".join(string.split()[-1:])
	book1["B28"].value = " ".join(string.split()[:-1]).capitalize() + " руб ,00к"


	output_book.save(output_location)
	output_book.close()



	messagebox.showinfo("Excel mover", "Данные перенесены!")
	quit()

#Properties
main_window.title('Excel mover')

main_window.geometry('500x250')

main_window.resizable(width=False, height=False)

#Variables	
works = StringVar()
materials = StringVar()
start_input_cell = StringVar(main_window, value='15')
start_output_cell = 26
olabel = False
ilabel = False
input_location = ''
output_location = ''


openfile_button = Button(main_window, text = "Файл с данными...", command=lambda: locations('open')).place(x = 10, y = 10)
output_button = Button(main_window, text = "Сохранить в...", command=lambda: locations('save')).place(x = 18, y = 60)

works_area = Entry(main_window, width = 3,textvariable=works).place(x = 50, y = 150)
materials_area = Entry(main_window, width = 3,textvariable=materials).place(x = 234, y = 150)
start_input_area = Entry(main_window, width = 3,textvariable=start_input_cell).place(x = 420, y = 150)

works_label =  Label(main_window, text='Работ:').place(x = 40, y = 120)
materials_label =  Label(main_window, text='Материалов').place(x = 208, y = 120)
start_input_label =  Label(main_window, text='Начиная с').place(x = 400, y = 120)




btn = Button(main_window, text = "Перенести", command=main).place(x = 210, y = 200)

main_window.mainloop()